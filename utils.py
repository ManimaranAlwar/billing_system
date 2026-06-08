import barcode
from barcode.writer import ImageWriter
import qrcode
import os
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from io import BytesIO
from flask import send_file
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import uuid

def generate_barcode(product_code, product_name):
    """Generate barcode for product"""
    try:
        # Generate EAN13 barcode (needs 12 digits + checksum)
        # Use product code padded to 12 digits
        code = str(product_code).zfill(12)
        EAN = barcode.get_barcode_class('ean13')
        ean = EAN(code, writer=ImageWriter())
        
        filename = f"barcode_{product_code}.png"
        filepath = os.path.join('static/uploads/barcodes', filename)
        
        # Save barcode
        ean.save(filepath)
        
        return filename
    except Exception as e:
        print(f"Error generating barcode: {e}")
        return None

def generate_qr_code(data):
    """Generate QR code for bill"""
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(data)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    
    filename = f"qr_{uuid.uuid4().hex[:8]}.png"
    filepath = os.path.join('static/uploads', filename)
    img.save(filepath)
    
    return filename

def generate_pdf_bill(order, items, shop_settings, customer=None, language='ta'):
    """Generate PDF bill"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=72)
    
    story = []
    styles = getSampleStyleSheet()
    is_english = language == 'en'
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#2ecc71'),
        alignment=1,
        spaceAfter=30
    )
    
    # Shop header
    story.append(Paragraph(shop_settings.shop_name_en if is_english else shop_settings.shop_name_ta, title_style))
    story.append(Paragraph(shop_settings.shop_name_en, styles['Heading2']))
    story.append(Paragraph(shop_settings.address_en if is_english else shop_settings.address_ta, styles['Normal']))
    story.append(Paragraph(f"{'Phone' if is_english else 'தொலைபேசி'}: {shop_settings.phone}", styles['Normal']))
    story.append(Paragraph(f"GST: {shop_settings.gst_no}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Bill details
    bill_info = [
        ['பில் எண்:', order.bill_number],
        ['தேதி:', order.order_date.strftime('%d/%m/%Y %H:%M')],
        ['வாடிக்கையாளர்:', customer.name if customer else 'பொது வாடிக்கையாளர்'],
        ['தொலைபேசி:', customer.phone if customer else '-'],
    ]
    if is_english:
        bill_info = [
            ['Bill No:', order.bill_number],
            ['Date:', order.order_date.strftime('%d/%m/%Y %H:%M')],
            ['Customer:', customer.name if customer else 'General Customer'],
            ['Phone:', customer.phone if customer else '-'],
        ]
    
    bill_table = Table(bill_info, colWidths=[2*inch, 3*inch])
    bill_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    story.append(bill_table)
    story.append(Spacer(1, 20))
    
    # Items table
    item_data = [['வ.எண்', 'பொருள்', 'அளவு', 'விலை (₹)', 'மொத்தம் (₹)']]
    if is_english:
        item_data[0] = ['S.No', 'Product', 'Qty', 'Price (Rs)', 'Total (Rs)']
    for idx, item in enumerate(items, 1):
        item_data.append([
            str(idx),
            item.product.name_en if is_english else item.product.name_ta,
            f"{item.quantity} {item.product.unit}",
            f"{item.price:.2f}",
            f"{item.total:.2f}"
        ])
    
    item_table = Table(item_data, colWidths=[0.5*inch, 2.5*inch, 1*inch, 1*inch, 1.2*inch])
    item_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(item_table)
    story.append(Spacer(1, 20))
    
    # Totals
    total_info = [
        ['மொத்தம்:', f"₹ {order.total_amount:.2f}"],
        ['தள்ளுபடி:', f"₹ {order.discount:.2f}"],
        ['வரி:', f"₹ {order.tax_amount:.2f}"],
        ['இறுதி மொத்தம்:', f"₹ {order.grand_total:.2f}"],
    ]
    if is_english:
        total_info = [
            ['Sub Total:', f"Rs {order.total_amount:.2f}"],
            ['Discount:', f"Rs {order.discount:.2f}"],
            ['Tax:', f"Rs {order.tax_amount:.2f}"],
            ['Grand Total:', f"Rs {order.grand_total:.2f}"],
        ]
    
    total_table = Table(total_info, colWidths=[3*inch, 2*inch])
    total_table.setStyle(TableStyle([
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TEXTCOLOR', (3, 3), (3, 3), colors.red),
    ]))
    story.append(total_table)
    story.append(Spacer(1, 30))
    
    # Footer
    story.append(Paragraph(shop_settings.bill_footer_en if is_english else shop_settings.bill_footer_ta, styles['Normal']))
    story.append(Paragraph(shop_settings.bill_footer_en, styles['Normal']))
    
    doc.build(story)
    buffer.seek(0)
    
    return buffer

def export_to_excel(data, headers, filename, tamil_headers=True):
    """Export data to Excel with Tamil headers"""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2ecc71", end_color="2ecc71", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write data
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer
